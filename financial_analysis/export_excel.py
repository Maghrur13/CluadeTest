"""
Export financial data and ratios to a formatted Excel workbook.
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


def _style_header(cell):
    cell.font = Font(bold=True, size=11)
    cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    cell.font = Font(bold=True, size=11, color="FFFFFF")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _style_section(cell):
    cell.font = Font(bold=True, size=10)
    cell.fill = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")


def _style_border(ws, start_row, end_row, start_col, end_col):
    thin = Side(style="thin", color="000000")
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            ws.cell(row=row, column=col).border = Border(
                left=thin, right=thin, top=thin, bottom=thin
            )


def export_to_excel(
    raw_data: dict,
    ratios: dict,
    output_path: str | Path,
    company_name: str = "Company",
) -> Path:
    """
    Create a formatted Excel workbook with:
    - Sheet 1: Extracted Data (balance sheet & income statement items)
    - Sheet 2: Financial Analysis (liquidity, solvency, profitability)
    """
    output_path = Path(output_path)
    wb = Workbook()

    # --- Sheet 1: Extracted Data ---
    ws1 = wb.active
    ws1.title = "Extracted Data"

    ws1["A1"] = f"Financial Statement Data - {company_name}"
    ws1["A1"].font = Font(bold=True, size=14)
    ws1.merge_cells("A1:B1")

    row = 3
    for key, val in raw_data.items():
        if key.startswith("_"):
            continue
        label = key.replace("_", " ").title()
        ws1.cell(row=row, column=1, value=label)
        ws1.cell(row=row, column=2, value=val if isinstance(val, (int, float)) else str(val))
        row += 1

    ws1.column_dimensions["A"].width = 28
    ws1.column_dimensions["B"].width = 18

    # --- Sheet 2: Financial Analysis ---
    ws2 = wb.create_sheet("Financial Analysis", 1)

    ws2["A1"] = f"Financial Statement Analysis - {company_name}"
    ws2["A1"].font = Font(bold=True, size=14)
    ws2.merge_cells("A1:C1")

    ws2["A2"] = "Ratio"
    ws2["B2"] = "Value"
    ws2["C2"] = "Interpretation"
    for c in ["A2", "B2", "C2"]:
        _style_header(ws2[c])

    row = 3
    interpretations = {
        "Current Ratio": ">1.5 healthy, <1 may indicate liquidity risk",
        "Quick Ratio (Acid Test)": ">1 preferred; excludes inventory",
        "Cash Ratio": "Most conservative liquidity measure",
        "Debt-to-Equity": "Lower = less leverage; varies by industry",
        "Debt-to-Assets": "Lower = less debt-financed",
        "Interest Coverage": ">2 healthy; ability to pay interest",
        "Net Profit Margin (%)": "Higher = more profitable per dollar of sales",
        "Return on Assets (ROA) (%)": "Efficiency of asset use",
        "Return on Equity (ROE) (%)": "Return to shareholders",
        "Gross Margin (%)": "Profit after direct costs",
        "Operating Margin (%)": "Profit from core operations",
    }

    for category, items in ratios.items():
        ws2.cell(row=row, column=1, value=category)
        _style_section(ws2.cell(row=row, column=1))
        ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        row += 1

        for ratio_name, value in items.items():
            ws2.cell(row=row, column=1, value=ratio_name)
            ws2.cell(row=row, column=2, value=value)
            ws2.cell(row=row, column=3, value=interpretations.get(ratio_name, ""))
            if isinstance(value, (int, float)) and value != 0:
                ws2.cell(row=row, column=2).number_format = "0.00"
            row += 1
        row += 1  # blank line between categories

    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 45

    wb.save(output_path)
    return output_path


def export_multi_to_excel(
    companies_data: dict[str, tuple[dict, dict]],
    output_path: str | Path,
) -> Path:
    """
    Export multiple companies to one Excel file.
    companies_data: {company_name: (raw_data, ratios)}
    """
    output_path = Path(output_path)
    wb = Workbook()
    ws = wb.active
    ws.title = "Multi-Company Comparison"

    ws["A1"] = "Financial Ratio Analysis - Multi-Company Comparison"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:" + get_column_letter(1 + len(companies_data)) + "1")

    # Headers: Ratio | Company1 | Company2 | ...
    ws.cell(row=2, column=1, value="Ratio")
    _style_header(ws.cell(row=2, column=1))
    for col, company_name in enumerate(companies_data, start=2):
        ws.cell(row=2, column=col, value=company_name)
        _style_header(ws.cell(row=2, column=col))

    row = 3
    for company_name, (_, ratios) in list(companies_data.items())[:1]:
        for category, items in ratios.items():
            ws.cell(row=row, column=1, value=category)
            _style_section(ws.cell(row=row, column=1))
            row += 1
            for ratio_name, value in items.items():
                ws.cell(row=row, column=1, value=ratio_name)
                row += 1
            row += 1

    # Fill values per company
    for col, company_name in enumerate(companies_data, start=2):
        _, ratios = companies_data[company_name]
        r = 3
        for category, items in ratios.items():
            r += 1
            for ratio_name, value in items.items():
                ws.cell(row=r, column=col, value=value)
                if isinstance(value, (int, float)):
                    ws.cell(row=r, column=col).number_format = "0.00"
                r += 1
            r += 1

    for c in range(1, len(companies_data) + 2):
        ws.column_dimensions[get_column_letter(c)].width = 20
    wb.save(output_path)
    return output_path
